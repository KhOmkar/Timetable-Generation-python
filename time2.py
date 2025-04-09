import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os
import re

class TimetableGenerator:
    def __init__(self):
        self.time_slots = [
            '8:30 to 9:25', '9:25 to 10:20', '10:20 to 10:30', '10:30 to 11:25',
            '11:25 to 12:20', '12:20 to 13:15', '13:15 to 14:10', '14:10 to 15:05',
            '15:05 to 15:10', '15:10 to 16:00', '16:00 to 16:50', '16:50 to 16:55',
            '16:55 to 17:45', '17:45 to 18:25'
        ]
        self.days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']

    def create_timetable_structure(self):
        df = pd.DataFrame(index=self.days, columns=self.time_slots)
        return df.fillna('')

    def process_all_sheets(self, input_file, classroom):
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Input file not found: {input_file}")

        try:
            excel_file = pd.ExcelFile(input_file)
            print(f"Found {len(excel_file.sheet_names)} sheets in the workbook")
            combined_schedule = self.create_timetable_structure()

            for sheet_name in excel_file.sheet_names:
                print(f"Processing sheet: {sheet_name}")
                workbook = load_workbook(input_file, data_only=True)
                sheet = workbook[sheet_name]
                division = sheet['N3'].value or f"Division ({sheet_name})"

                raw_timetable = pd.read_excel(input_file, 
                                           sheet_name=sheet_name,
                                           skiprows=6, 
                                           nrows=25)

                for index, row in raw_timetable.iterrows():
                    day = str(row.iloc[0]).strip()
                    if day not in self.days:
                        continue

                    for col_idx, time_slot in enumerate(self.time_slots):
                        current_cell = str(row.iloc[col_idx + 1])
                        if self.is_classroom_in_cell(current_cell, classroom):
                            components = current_cell.strip().split()
                            cell_content = "\n".join([
                                " ".join(components[:2]),
                                " ".join(components[2:]),
                                f"({division})"
                            ])
                            
                            existing_content = combined_schedule.at[day, time_slot]
                            if existing_content:
                                cell_content = f"{existing_content}\n---\n{cell_content}"
                            
                            combined_schedule.at[day, time_slot] = cell_content

            return combined_schedule

        except Exception as e:
            raise Exception(f"Error processing sheets: {str(e)}")

    def is_classroom_in_cell(self, cell_content, target_classroom):
        classrooms = re.findall(r'H[A-Z]?\d+[A-Z]?', cell_content.upper())
        return any(target_classroom.upper() == cls for cls in classrooms)

    def save_classroom_schedule(self, schedule_df, output_file, classroom):
        """
        Saves the processed schedule to an Excel file with proper formatting.
        Includes styling, cell alignment, and automatic size adjustments.
        Adds metadata section below the timetable.
        """
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Save the DataFrame to Excel
                schedule_df.to_excel(writer, sheet_name=f'Schedule_{classroom}')
                workbook = writer.book
                worksheet = writer.sheets[f'Schedule_{classroom}']

                # Add a title row at the top
                worksheet.insert_rows(1)
                title_cell = worksheet['B1']
                title_cell.value = f"Combined Classroom Schedule - {classroom}"
                title_cell.font = Font(bold=True, size=14)

                # Format all cells in the worksheet
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        # Set alignment for all cells
                        cell.alignment = Alignment(wrap_text=True, 
                                                vertical='center', 
                                                horizontal='center')
                        
                        # Highlight headers (first row and first column)
                        if cell.row == 2 or cell.column == 1:
                            cell.fill = PatternFill(start_color="E0E0E0", 
                                                end_color="E0E0E0", 
                                                fill_type="solid")
                            cell.font = Font(bold=True)

                # Adjust column widths based on content
                max_width = 0
                for column in worksheet.columns:
                    for cell in column:
                        if cell.value:
                            max_width = max(max_width, len(str(cell.value)))

                uniform_width = min(max_width // 2, 25) 
                for column in worksheet.columns:
                    worksheet.column_dimensions[column[0].column_letter].width = uniform_width

                # Adjust row heights based on content
                for row in worksheet.rows:
                    max_lines = max(str(cell.value).count('\n') + 1 if cell.value else 1 for cell in row)
                    worksheet.row_dimensions[row[0].row].height = max_lines * 15

                # Handle merged cells
                for col_idx, column in enumerate(worksheet.iter_cols(), start=1):
                    for row_idx in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value and len(str(cell.value)) > 20:
                            next_col_idx = col_idx + 1
                            if next_col_idx <= worksheet.max_column:
                                next_cell = worksheet.cell(row=row_idx, column=next_col_idx)
                                worksheet.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=next_col_idx)
                                next_cell.value = None

                # Merge rows 3 to 8 for selected time slot columns
                selected_slots = ['10:20 to 10:30', '12:20 to 13:15', '15:05 to 15:10', '16:50 to 16:55']
                slot_columns = [col for col in schedule_df.columns if col in selected_slots]

                # Get width of first column to apply to selected columns
                first_col_letter = get_column_letter(1)  # Column B (first time slot)
                first_col_width = worksheet.column_dimensions[first_col_letter].width

                break_labels = ["SHORT BREAK 1", "LUNCH BREAK", "SHORT BREAK 2", "SHORT BREAK 3"]

                for slot, label in zip(slot_columns, break_labels):
                    try:
                        col_idx = list(schedule_df.columns).index(slot) + 2  # +2 because A is index, B is first time slot
                        col_letter = get_column_letter(col_idx)

                        # Merge rows 3 to 8 in this column
                        worksheet.merge_cells(start_row=3, end_row=8, start_column=col_idx, end_column=col_idx)

                        # Set the column width same as the first column
                        worksheet.column_dimensions[col_letter].width = first_col_width

                        # Add the break label vertically in the merged cell
                        merged_cell = worksheet.cell(row=3, column=col_idx)
                        merged_cell.value = label
                        merged_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
                        merged_cell.font = Font(bold=True)

                    except Exception as merge_err:
                        print(f"Warning: Could not merge column '{slot}' - {merge_err}")

                # Add metadata section below the timetable
                self._add_metadata_section(worksheet, classroom)

                # Save the final workbook
                workbook.save(output_file)

        except Exception as e:
            raise Exception(f"Error saving schedule: {str(e)}")

    def _add_metadata_section(self, worksheet, classroom):
        """Adds metadata section with merged cells and combined divisions"""
        try:
            # Find the last row of the timetable
            last_row = worksheet.max_row
            
            # Skip 2 rows after the timetable
            metadata_start_row = last_row + 2
            
            # Add metadata header
            headers = ["Course Code", "Course Name", "Teacher Name", "Divisions"]
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=metadata_start_row, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9EAD3", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Set double height for header row
            worksheet.row_dimensions[metadata_start_row].height = 30
            
            # Try to load metadata files
            meta_files = [
                "C:/Users/omkar/Downloads/timetable/meta_info_Theory_section.csv",
                "C:/Users/omkar/Downloads/timetable/meta_info_Practical_section.csv"
            ]
            
            metadata_entries = []
            
            for file in meta_files:
                if os.path.exists(file):
                    df = pd.read_csv(file)
                    # Filter for current classroom
                    classroom_filter = df['Classroom'].str.contains(classroom, na=False)
                    filtered_df = df[classroom_filter]
                    
                    # Collect all metadata entries
                    for _, row in filtered_df.iterrows():
                        metadata_entries.append({
                            'Course_Code': row.get('Course_Code', ''),
                            'Course_Name': row.get('Course_Name', ''),
                            'Initials': row.get('Course_Initials', ''),
                            'Teacher_Name': row['Teacher_Name'],
                            'Teacher_Initials': row.get('Teacher_Initials', ''),
                            'Division': row.get('Division', '')
                        })
            
            if not metadata_entries:
                return
                
            # Convert to DataFrame for grouping
            meta_df = pd.DataFrame(metadata_entries)
            
            # Group by course code, name, and teacher
            grouped = meta_df.groupby(['Course_Code', 'Course_Name', 'Teacher_Name', 'Teacher_Initials'])
            
            # Combine divisions for each group
            combined_meta = grouped.agg({
                'Division': lambda x: ', '.join(sorted(set(x))),
                'Initials': 'first'
            }).reset_index()
            
            current_row = metadata_start_row + 1
            
            # Group again just by course code for merging
            course_groups = combined_meta.groupby(['Course_Code', 'Course_Name'])
            
            for (course_code, course_name), course_group in course_groups:
                row_span = len(course_group)
                
                # Write course code and name (will be merged vertically)
                worksheet.cell(row=current_row, column=1, value=course_code)
                worksheet.cell(row=current_row, column=2, value=f"{course_name} ({course_group.iloc[0]['Initials']})")
                
                # Merge course code and name cells vertically if multiple teachers
                if row_span > 1:
                    worksheet.merge_cells(
                        start_row=current_row,
                        end_row=current_row + row_span - 1,
                        start_column=1,
                        end_column=1
                    )
                    worksheet.merge_cells(
                        start_row=current_row,
                        end_row=current_row + row_span - 1,
                        start_column=2,
                        end_column=2
                    )
                
                # Write teacher info and combined divisions
                for _, row in course_group.iterrows():
                    teacher_display = f"{row['Teacher_Name']} ({row['Teacher_Initials']})" if row['Teacher_Initials'] else row['Teacher_Name']
                    
                    worksheet.cell(row=current_row, column=3, value=teacher_display)
                    worksheet.cell(row=current_row, column=4, value=row['Division'])
                    
                    # Set double height and alignment
                    worksheet.row_dimensions[current_row].height = 30
                    for col in range(1, 5):
                        worksheet.cell(row=current_row, column=col).alignment = Alignment(
                            horizontal='center', 
                            vertical='center',
                            wrap_text=True
                        )
                    
                    current_row += 1
            
            # Adjust column widths
            for col_idx in range(1, len(headers) + 1):
                max_length = 0
                for row in worksheet.iter_rows(min_row=metadata_start_row, max_row=current_row-1, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
                
        except Exception as e:
            print(f"Warning: Could not add metadata section - {str(e)}")

def main():
    input_file = "D:\\Classwise 24 25 Sem I.xlsm"
    output_file = "C:\\Users\\omkar\\Downloads\\timetable\\Classroom_Schedule_SBK.xlsx"
    classroom = "SBK"

    try:
        generator = TimetableGenerator()
        print(f"Generating schedule for classroom {classroom}...")
        schedule = generator.process_all_sheets(input_file, classroom)
        generator.save_classroom_schedule(schedule, output_file, classroom)
        print(f"Schedule with metadata saved to {output_file}")

    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()