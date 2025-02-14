import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

class TimetableGenerator:
    def __init__(self):
        self.time_slots = [
            '8:30 to 9:25', '9:25 to 10:20', '10:20 to 10:30', '10:30 to 11:25',
            '11:25 to 12:20', '12:20 to 13:15', '13:15 to 14:10', '14:10 to 15:05',
            '15:05 to 15:10', '15:10 to 16:00', '16:00 to 16:50', '16:50 to 16:55',
            '16:55 to 17:45', '17:45 to 18:25'
        ]
        self.days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']

    def create_timetable_structure(self):
        """Creates an empty timetable DataFrame with days and time slots."""
        df = pd.DataFrame(index=self.days, columns=self.time_slots)
        return df.fillna('')

    def extract_classroom_schedule(self, input_file, classroom):
        """
        Extracts schedule for a specific classroom from the input Excel file.
        
        Args:
            input_file (str): Path to input Excel file
            classroom (str): Classroom code (e.g., "H303")
            
        Returns:
            pd.DataFrame: Processed classroom schedule
        """
        # Validate input file
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Input file not found: {input_file}")

        try:
            # Load workbook and get division
            workbook = load_workbook(input_file, data_only=True)
            sheet = workbook.active
            division = sheet['N3'].value if sheet['N3'].value else "Unknown Division"

            # Read timetable data
            raw_timetable = pd.read_excel(input_file, skiprows=6, nrows=25)
            classroom_schedule = self.create_timetable_structure()

            # Process each row in timetable
            for index, row in raw_timetable.iterrows():
                day = row.iloc[0]
                if not isinstance(day, str) or day.strip() not in self.days:
                    continue
                
                day = day.strip()
                
                # Process each time slot
                for col_idx, time_slot in enumerate(self.time_slots):
                    current_cell = row.iloc[col_idx + 1]
                    if pd.isna(current_cell) or not isinstance(current_cell, str):
                        continue
                    
                    # Extract classroom information
                    if classroom in current_cell:
                        components = current_cell.strip().split()
                        # Format cell content with subject code, faculty, and division
                        cell_content = "\n".join([
                            " ".join(components[:2]),  # Subject code
                            " ".join(components[2:]),  # Faculty and other info
                            f"({division})"            # Division
                        ])
                        classroom_schedule.at[day, time_slot] = cell_content

            return classroom_schedule

        except Exception as e:
            raise Exception(f"Error processing timetable: {str(e)}")

    def save_classroom_schedule(self, schedule_df, output_file, classroom):
        """
        Saves the processed schedule to an Excel file with formatting.
        
        Args:
            schedule_df (pd.DataFrame): Processed schedule
            output_file (str): Output file path
            classroom (str): Classroom code
        """
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Save DataFrame
                schedule_df.to_excel(writer, sheet_name=f'Schedule_{classroom}')
                workbook = writer.book
                worksheet = writer.sheets[f'Schedule_{classroom}']

                # Add title
                worksheet.insert_rows(1)
                title_cell = worksheet['B1']
                title_cell.value = f"Classroom Schedule - {classroom}"
                title_cell.font = Font(bold=True, size=14)

                # Format cells
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, 
                                                vertical='center', 
                                                horizontal='center')
                        
                        # Highlight headers
                        if cell.row == 2 or cell.column == 1:
                            cell.fill = PatternFill(start_color="E0E0E0", 
                                                  end_color="E0E0E0", 
                                                  fill_type="solid")
                            cell.font = Font(bold=True)

                # Adjust column widths and row heights
                for column in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in column if cell.value)
                    worksheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

                for row in worksheet.rows:
                    max_lines = max(str(cell.value).count('\n') + 1 if cell.value else 1 for cell in row)
                    worksheet.row_dimensions[row[0].row].height = max_lines * 15

        except Exception as e:
            raise Exception(f"Error saving schedule: {str(e)}")

def main():
    # File paths
    input_file = "D:\\Classwise 24 25 Sem I 05.xlsm"
    output_file = "C:\\Users\\omkar\\Downloads\\timetable\\H3032.xlsx"
    classroom = "H303"
    
    try:
        generator = TimetableGenerator()
        print(f"Extracting schedule for classroom {classroom}...")
        classroom_schedule = generator.extract_classroom_schedule(input_file, classroom)
        
        print("Saving schedule to Excel...")
        generator.save_classroom_schedule(classroom_schedule, output_file, classroom)
        
        print(f"Schedule has been generated and saved to {output_file}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()