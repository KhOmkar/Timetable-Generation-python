import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import os

class TimetableGenerator:
    def __init__(self):
        # Define all time slots for the timetable 
        # These represent different periods throughout the day
        self.time_slots = [
            '8:30 to 9:25', '9:25 to 10:20', '10:20 to 10:30', '10:30 to 11:25',
            '11:25 to 12:20', '12:20 to 13:15', '13:15 to 14:10', '14:10 to 15:05',
            '15:05 to 15:10', '15:10 to 16:00', '16:00 to 16:50', '16:50 to 16:55',
            '16:55 to 17:45', '17:45 to 18:25'
        ]
        # Define working days of the week
        self.days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']


    def create_timetable_structure(self):
        """
        Creates an empty timetable structure as a pandas DataFrame.
        - Rows represent days of the week
        - Columns represent time slots
        - All cells are initialized as empty strings
        """
        df = pd.DataFrame(index=self.days, columns=self.time_slots)
        # print(df)
        return df.fillna('')  # Fill NaN values with empty strings

    def process_all_sheets(self, input_file, classroom):
        """
        Processes all sheets in the Excel file to create a combined classroom schedule.
        Each sheet typically represents a different division's timetable.SS
        """
        # Check if input file exists
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"Input file not found: {input_file}")

        try:
            # Create Excel file object and get list of all sheet names
            excel_file = pd.ExcelFile(input_file)
            sheet_names = excel_file.sheet_names
            print(f"Found {len(sheet_names)} sheets in the workbook")

            # Create empty schedule DataFrame to store combined results
            combined_schedule = self.create_timetable_structure()

            # Process each sheet in the workbook
            for sheet_name in sheet_names:
                print(f"Processing sheet: {sheet_name}")
                
                # Load the current sheet from the workbook
                workbook = load_workbook(input_file, data_only=True)
                sheet = workbook[sheet_name]
               
                # Get division information from cell N3
                # If N3 is empty, use sheet name as division identifier
                division = sheet['N3'].value if sheet['N3'].value else f"Division ({sheet_name})"

                # Read timetable data from current sheet
                # Skip first 6 rows (header information)
                # Read only 25 rows (timetable content)
                raw_timetable = pd.read_excel(input_file, 
                                            sheet_name=sheet_name,
                                            skiprows=6, 
                                            nrows=25)

                # Process each row (day) in the timetable
                for index, row in raw_timetable.iterrows():
                    day = row.iloc[0]  # First column contains day information
                    
                    # Skip if day is not valid
                    if not isinstance(day, str) or day.strip() not in self.days:
                        continue
                    
                    day = day.strip()
                    
                    # Process each time slot in the current day
                    for col_idx, time_slot in enumerate(self.time_slots):
                        # Get cell content (add 1 to col_idx because first column is day)
                        current_cell = row.iloc[col_idx + 1]
                        
                        # Skip empty or non-string cells
                        if pd.isna(current_cell) or not isinstance(current_cell, str):
                            continue
                        
                        # Check if this time slot involves the classroom we're interested in
                        if classroom in current_cell:
                            # Split cell content into components
                            components = current_cell.strip().split()
                            #   print(current_cell)
                            # Format cell content in three lines:
                            # 1. Subject code
                            # 2. Faculty and other information
                            # 3. Division information
                            cell_content = "\n".join([
                                " ".join(components[:2]),     # First two components are subject code
                                " ".join(components[2:]),     # Remaining components are faculty info
                                f"({division})"              # Division information in parentheses
                            ])
                            
                            # If this time slot already has content, append new content
                            # Use "---" as a separator between different classes
                            existing_content = combined_schedule.at[day, time_slot]
                            if existing_content:
                                cell_content = f"{existing_content}\n---\n{cell_content}"
                            
                            # Update the schedule with the new content
                            combined_schedule.at[day, time_slot] = cell_content

            return combined_schedule

        except Exception as e:
            raise Exception(f"Error processing sheets: {str(e)}")

    def save_classroom_schedule(self, schedule_df, output_file, classroom):
        """
        Saves the processed schedule to an Excel file with proper formatting.
        Includes styling, cell alignment, and automatic size adjustments.
        """
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Save the DataFrame to Excel
                schedule_df.to_excel(writer, sheet_name=f'Schedule_{classroom}')
                workbook = writer.book
                worksheet = writer.sheets[f'Schedule_{classroom}']

                # Add a title row at the top
                worksheet.insert_rows(1)
                title_cell = worksheet['B1']
                title_cell.value = f"Combined Classroom Schedule - {classroom}"
                title_cell.font = Font(bold=True, size=14)

                # Format all cells in the worksheet
                for row in worksheet.iter_rows(min_row=2):
                    for cell in row:
                        # Set alignment for all cells
                        cell.alignment = Alignment(wrap_text=True, 
                                                vertical='center', 
                                                horizontal='center')
                        
                        # Highlight headers (first row and first column)
                        if cell.row == 2 or cell.column == 1:
                            cell.fill = PatternFill(start_color="E0E0E0", 
                                                  end_color="E0E0E0", 
                                                  fill_type="solid")
                            cell.font = Font(bold=True)

                # Adjust column widths based on content
                # Calculate the maximum column width among all columns
                max_width = 0
                for column in worksheet.columns:
                    for cell in column:
                        if cell.value:
                            max_width = max(max_width, len(str(cell.value)))

                # Set all columns to half of the max width, ensuring it's within a reasonable range
                uniform_width = min(max_width // 2, 25)  # Half of max width, with an upper limit of 25

                # Apply uniform width to all columns
                for column in worksheet.columns:
                    worksheet.column_dimensions[column[0].column_letter].width = uniform_width


                # Adjust row heights based on content (especially for cells with multiple lines)
                for row in worksheet.rows:
                    max_lines = max(str(cell.value).count('\n') + 1 if cell.value else 1 for cell in row)
                    worksheet.row_dimensions[row[0].row].height = max_lines * 15

                for col_idx, column in enumerate(worksheet.iter_cols(), start=1):
                    for row_idx in range(2, worksheet.max_row + 1):  # Skip title row
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if cell.value and len(str(cell.value)) > 20:
                            next_col_idx = col_idx + 1  # Next column index
                            if next_col_idx <= worksheet.max_column:  # Ensure within range
                                next_cell = worksheet.cell(row=row_idx, column=next_col_idx)
                                worksheet.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx, end_column=next_col_idx)
                                next_cell.value = None  # Clear merged cell

            # Save the final workbook
            workbook.save(output_file)

        except Exception as e:
            raise Exception(f"Error saving schedule: {str(e)}")

def main():
    # Define file paths for input and output
    input_file = "D:\\Classwise 24 25 Sem I.xlsm"
    output_file = "C:\\Users\\omkar\\Downloads\\timetable\\H202_column_join-size.xlsx"
    classroom = "H202"
    
    try:
        # Create timetable generator instance
        generator = TimetableGenerator()
        print(f"Processing all sheets for classroom {classroom}...")
        
        # Process all sheets and combine schedules
        combined_schedule = generator.process_all_sheets(input_file, classroom)
        
        # Save the combined schedule to Excel
        print("Saving combined schedule to Excel...")
        generator.save_classroom_schedule(combined_schedule, output_file, classroom)
        
        (f"Combined schedule has been generated and saved to {output_file}")
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()