import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def create_timetable_structure():
    """
    Creates the basic structure of the timetable with predefined time slots and days.
    Returns a pandas DataFrame with days as index and time slots as columns.
    """
    # Define all possible time slots in the timetable
    # These correspond to columns B through O in the Excel file
    time_slots = [
        '8:30 to 9:25',    # Column B - First period
        '9:25 to 10:20',   # Column C - Second period
        '10:20 to 10:30',  # Column D - First break
        '10:30 to 11:25',  # Column E - Third period
        '11:25 to 12:20',  # Column F - Fourth period
        '12:20 to 13:15',  # Column G - Fifth period
        '13:15 to 14:10',  # Column H - Sixth period
        '14:10 to 15:05',  # Column I - Seventh period
        '15:05 to 15:10',  # Column J - Second break
        '15:10 to 16:00',  # Column K - Eighth period
        '16:00 to 16:50',  # Column L - Ninth period
        '16:50 to 16:55',  # Column M - Third break
        '16:55 to 17:45',  # Column N - Tenth period
        '17:45 to 18:25'   # Column O - Eleventh period
    ]
    
    # Define working days
    days = ['MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT']
    
    # Create empty DataFrame and fill NaN values with empty strings
    df = pd.DataFrame(index=days, columns=time_slots)
    return df.fillna('')

def is_break_slot(time_slot):
    """
    Determines if a given time slot is a break period.
    
    Args:
        time_slot (str): The time slot to check
    
    Returns:
        bool: True if it's a break slot, False otherwise
    """
    # List of times that indicate breaks in the schedule
    break_times = ['10:20', '15:05', '16:50']
    return any(break_time in time_slot for break_time in break_times)

def extract_teacher_schedule(input_file, teacher_short="MNV"):
    """
    Extracts a teacher's schedule from the input Excel file.
    
    Args:
        input_file (str): Path to the input Excel file
        teacher_short (str): Teacher's short code to search for (e.g., "MNV")
    
    Returns:
        pandas.DataFrame: Extracted schedule with merged cells indicated
    """
    # Read Excel file, skipping header rows
    raw_timetable = pd.read_excel(input_file, skiprows=6, nrows=25)
   # print (raw_timetable)
    # Create empty timetable structure
    teacher_schedule = create_timetable_structure()
    time_slots = teacher_schedule.columns.tolist()
    raw_timetable.to_excel("C:\\Users\\omkar\\Downloads\\timetableraw_timetable_complete.xlsx", index=False)
    # Process each row in the timetable
    for index, row in raw_timetable.iterrows():
        day = row.iloc[0]  # First column contains the day
        
        # Skip invalid or empty day entries
        if not isinstance(day, str) or day.strip() not in teacher_schedule.index:
            continue
        
        day = day.strip()
        skip_next = False  # Flag to skip next column if it's part of a practical
        
        # Process each time slot in the row
        for col_idx, time_slot in enumerate(time_slots):
            # Skip if this slot is second part of a practical
            if skip_next:
                skip_next = False
                continue
                
            current_cell = row.iloc[col_idx + 1]  # +1 because first column is day
            
            # Skip empty or non-string cells
            if pd.isna(current_cell) or not isinstance(current_cell, str):
                continue
                
            # Process cells containing teacher's code
            if teacher_short in current_cell:
                # Check if this is a practical class (spans two columns)
                is_practical = False
                
                # Look ahead to next column if it exists
                if col_idx < len(time_slots) - 1:
                    next_slot = time_slots[col_idx + 1]
                    next_cell = row.iloc[col_idx + 2]  # +2 because first column is day
                    
                    # Practical class conditions:
                    # 1. Next slot is not a break
                    # 2. Next cell contains same teacher code
                    if (not is_break_slot(next_slot) and 
                        isinstance(next_cell, str) and 
                        teacher_short in next_cell):
                        is_practical = True
                
                # Clean up cell content by removing extra whitespace
                current_content = '\n'.join(line.strip() for line in current_cell.split('\n') if line.strip())
                
                if is_practical:
                    # For practical classes: use two columns
                    teacher_schedule.at[day, time_slot] = current_content
                    teacher_schedule.at[day, next_slot] = "MERGED_CELL"  # Marker for cells to be merged
                    skip_next = True  # Skip processing next column
                else:
                    # For theory classes: use single column
                    teacher_schedule.at[day, time_slot] = current_content
    
    return teacher_schedule

def save_teacher_schedule(schedule_df, output_file, teacher_short="MNV"):
    """
    Saves the schedule to an Excel file with proper formatting.
    
    Args:
        schedule_df (pandas.DataFrame): The schedule to save
        output_file (str): Output file path
        teacher_short (str): Teacher's code for sheet naming
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write DataFrame to Excel
        schedule_df.to_excel(writer, sheet_name=f'Schedule_{teacher_short}')
        
        workbook = writer.book
        worksheet = writer.sheets[f'Schedule_{teacher_short}']
        
        # Adjust column widths based on content
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value and cell.value != "MERGED_CELL":
                        # Calculate maximum length considering line breaks
                        cell_lines = str(cell.value).split('\n')
                        line_lengths = [len(line) for line in cell_lines]
                        if line_lengths:
                            if max(line_lengths) > max_length:
                                max_length = max(line_lengths)
                except:
                    pass
            # Set column width (max 50 characters)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Process cells for merging and formatting
        for row_idx in range(2, worksheet.max_row + 1):  # Start from 2 to skip header
            for col_idx in range(2, worksheet.max_column + 1):  # Start from 2 to skip index
                cell = worksheet.cell(row=row_idx, column=col_idx)
                
                if cell.value == "MERGED_CELL":
                    # Clear the marker text
                    cell.value = None
                    
                    # Merge with previous cell
                    prev_cell = worksheet.cell(row=row_idx, column=col_idx - 1)
                    merge_range = f"{prev_cell.column_letter}{row_idx}:{cell.column_letter}{row_idx}"
                    worksheet.merge_cells(merge_range)
                    
                    # Format merged cell
                    prev_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    # Format regular cells
                    if cell.value:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Adjust row heights based on content
        for row in worksheet.rows:
            max_lines = 1
            for cell in row:
                if cell.value and cell.value != "MERGED_CELL":
                    lines = str(cell.value).count('\n') + 1
                    if lines > max_lines:
                        max_lines = lines
            # Set row height (15 points per line)
            worksheet.row_dimensions[row[0].row].height = max_lines * 15

def main():
    # File paths configuration
    input_file = "D:\\Classwise 24 25 Sem I 05.xlsm"
    output_file = "C:\\Users\\omkar\\Downloads\\timetable\\mnv5.xlsx"
    
    try:
        # Extract and save schedule
        print(f"Extracting schedule for teacher MNV...")
        teacher_schedule = extract_teacher_schedule(input_file, "MNV")
        
        print("Saving schedule to Excel...")
        save_teacher_schedule(teacher_schedule, output_file)
        
        print(f"Schedule has been generated and saved to {output_file}")
        
    except Exception as e:
        print(f"An error occurred: {str(e)}")

    

if __name__ == "__main__":
    main()